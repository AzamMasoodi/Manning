'''

a:The frontal area per unit volume parameter (m-1), for channel it is about 100
C: coefficient to parameteriz the shear stress at the interface between vegetated and unvegetated regions.
    The range of this parameter is 0.05–0.13 (Luhar and Nepf, 2012).
kLN: a constant to correct the dimension of equation (m1/3.s-1)
Bx:Blockage factor
DensVeg=(m*(np.pi)*dVeg^2)/4

if flexibility==0:
'''
import openpyxl 
import matplotlib.pyplot as plt
grav_sqrt=pow(9.81, 0.5)

def Manning(WHr, PH, wVeg, FW, C, CDrag, a, kLN):
    """
    Calcualtates Manning n accdording to Luhar and Nepf(2012) doi: 

    Parameters
    ----------
    WHr : float
        Water surface depth in m
    PH : float
        Height of Vegation in m
    wVeg : float
        Width of Vegation in m
    FW : float
        Width of flow in m
    C : float
        coefficient to parameteriz the shear stress at the interface between vegetated and unvegetated regions.
            The range of this parameter is 0.05–0.13
    CDrag : float
        DESCRIPTION.
    a : float
        The frontal area per unit volume parameter (m-1), for channel it is about 100
    kLN : float
        a constant to correct the dimension of equation in :math:`m^{1/3}*s^{-1}`


    Returns
    -------
    float
        Manning n value for current flow and vegetation conditions in :math:`m^{1/3}*s^{-1}`

    """
    if WHr<=PH:
        Bx = wVeg/FW
        if Bx < 0.8:
            #  Eq.24
            return ((kLN*(WHr**(1/6)))/grav_sqrt)*((C/2)**0.5)*((1-Bx)**(-3/2)), Bx
        else:
            #  Eq.26
            return ((kLN*(WHr**(1/6)))/grav_sqrt)*((CDrag*a*WHr/2)**(1/2)), Bx
    else:
        #HWr>PH
        Bx = (wVeg*PH)/(FW*WHr)
        #  Eq.29 pp.14
        return ((kLN*(WHr**(1/6)))/grav_sqrt)*(1/(((2/C)**0.5)*((1-PH/WHr)**1.5)+((2*PH/(CDrag*a))**0.5)*(1/WHr))), Bx

def ManningLoop(CDrag, a, C, kLN):

    """
    

    Parameters
    ----------
    CDrag : float
        Drag coefficient of vegetation
    a : float
        The frontal area per unit volume parameter (m-1), for channel it is about 100
    C : TYPE
        coefficient to parameteriz the shear stress at the interface between vegetated and unvegetated regions.
    kLN : a constant to correct the dimension of equation in :math:`m^{1/3}*s^{-1}`

    Returns
    -------
    list_n : Manning n
    list_Bx : Vegetation Blockage
    list_DWH : Depth of water (m)
    list_Kstr : Strikler Coefficient

    """
    wrkbk = openpyxl.load_workbook("Data.xlsx") 
    df = wrkbk.active
    list_n=[]
    list_Bx=[]
    list_WHr=[]
    list_Kstr=[]
    for i in range(2, df.max_row+1):
        WHr=(df.cell(row=i,column=1)).value
        FW=(df.cell(row=i,column=2)).value
        PH=(df.cell(row=i,column=3)).value
        wVeg=(df.cell(row=i,column=4)).value
        n, Bx = Manning(WHr, PH, wVeg, FW, C, CDrag, a, kLN)
        list_n.append(n)
        list_Bx.append(Bx)
        list_WHr.append(WHr)
        list_Kstr.append(1/n)
    return list_n, list_Bx, list_WHr, list_Kstr

list_n, list_Bx, list_WHr, list_Kstr = ManningLoop( 1, 100, 0.052, 1)

plt.plot(list_Bx, list_n, 'ro')
plt.xlabel('Block index')
plt.ylabel('Manning coeficient')
plt.show()
plt.plot(list_Kstr, list_WHr)
plt.xlabel('K_str')
plt.ylabel('Depth of water(m)')
plt.show()
#print('n:',["%.3f" % elem for elem in list_n])
#print('Bx:',["%.3f" % elem for elem in list_Bx])

